from __future__ import annotations

import argparse
import csv
import gzip
import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from scipy.stats import norm
from statsmodels.duration.hazard_regression import PHReg


COEFFICIENTS = {
    "TPI1": 0.3041908,
    "ENO1": 0.9639654,
    "LDHA": 1.3404374,
    "SLC2A1": 0.2424239,
}


def read_geo_table(path: Path, begin_marker: str, end_marker: str) -> pd.DataFrame:
    rows: list[list[str]] = []
    in_table = False
    with gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        for line in handle:
            stripped = line.rstrip("\r\n")
            if stripped == begin_marker:
                in_table = True
                continue
            if stripped == end_marker:
                break
            if in_table:
                rows.append(next(csv.reader([stripped], delimiter="\t", quotechar='"')))
    if not rows:
        raise ValueError(f"No table found in {path}")
    header = rows[0]
    return pd.DataFrame(rows[1:], columns=header)


def parse_annotation(path: Path) -> pd.DataFrame:
    with gzip.open(path, "rt", encoding="utf-8", errors="replace") as handle:
        lines = handle.readlines()
    header_idx = next(i for i, line in enumerate(lines) if line.startswith("ID\t"))
    text = "".join(lines[header_idx:])
    from io import StringIO

    annotation = pd.read_csv(StringIO(text), sep="\t", dtype=str)
    symbol_col = next(
        col
        for col in annotation.columns
        if col.lower().replace(" ", "") in {"genesymbol", "symbol"}
    )
    annotation = annotation[["ID", symbol_col]].rename(columns={symbol_col: "gene_symbol"})
    annotation["gene_symbol"] = annotation["gene_symbol"].fillna("")
    return annotation


def extract_model_expression(matrix: pd.DataFrame, annotation: pd.DataFrame) -> pd.DataFrame:
    matrix = matrix.rename(columns={matrix.columns[0]: "ID"})
    matrix = matrix.merge(annotation, on="ID", how="inner")
    records = []
    for gene in COEFFICIENTS:
        gene_rows = matrix[
            matrix["gene_symbol"].str.split(r"\s*///\s*", regex=True).apply(
                lambda values: gene in values
            )
        ]
        if gene_rows.empty:
            raise ValueError(f"No annotated probes found for {gene}")
        values = gene_rows.drop(columns=["ID", "gene_symbol"]).apply(pd.to_numeric, errors="coerce")
        means = values.mean(axis=0)
        records.append(pd.DataFrame({"sample": means.index, gene: means.values}))
    out = records[0]
    for frame in records[1:]:
        out = out.merge(frame, on="sample", how="inner")
    return out


def parse_binary(series: pd.Series, positive: set[str], negative: set[str]) -> pd.Series:
    normalized = series.astype(str).str.strip().str.lower()
    out = pd.Series(np.nan, index=series.index, dtype=float)
    out[normalized.isin(positive)] = 1.0
    out[normalized.isin(negative)] = 0.0
    return out


def build_analysis_table(clinical_path: Path, expression: pd.DataFrame) -> pd.DataFrame:
    clinical = pd.read_csv(clinical_path, sep="\t", compression="gzip", dtype=str)
    out = pd.DataFrame(
        {
            "sample": clinical["Affy_GSM"],
            "tissue_type": clinical["Tissue Type"],
            "OS_status": pd.to_numeric(clinical["Survival status"], errors="coerce"),
            "OS_time_months": pd.to_numeric(clinical["Survival months"], errors="coerce"),
            "AFP_high": parse_binary(clinical["AFP (>/<=300ng/ml)"], {"high"}, {"low"}),
            "cirrhosis_yes": parse_binary(clinical["Cirrhosis"], {"y", "yes"}, {"n", "no"}),
            "size_large": parse_binary(
                clinical["Main Tumor Size (>/<=5 cm)"], {"large"}, {"small"}
            ),
            "multinodular_yes": parse_binary(
                clinical["Multinodular"], {"y", "yes"}, {"n", "no"}
            ),
        }
    )
    out = out[out["tissue_type"].str.lower().eq("tumor")].merge(expression, on="sample", how="inner")
    out["four_gene_score"] = sum(out[gene] * coef for gene, coef in COEFFICIENTS.items())
    out["score_z"] = (out["four_gene_score"] - out["four_gene_score"].mean()) / out[
        "four_gene_score"
    ].std(ddof=0)
    return out


def fit_cox(data: pd.DataFrame, model_name: str, covariates: list[str]) -> tuple[pd.DataFrame, dict]:
    required = ["OS_time_months", "OS_status", *covariates]
    fit_data = data.dropna(subset=required).copy()
    model = PHReg(
        fit_data["OS_time_months"].astype(float),
        fit_data[covariates].astype(float),
        status=fit_data["OS_status"].astype(int),
        ties="efron",
    )
    result = model.fit(disp=False)
    params = np.asarray(result.params)
    bse = np.asarray(result.bse)
    pvalues = np.asarray(result.pvalues)
    z = norm.ppf(0.975)
    terms = pd.DataFrame(
        {
            "model": model_name,
            "term": covariates,
            "coefficient": params,
            "standard_error": bse,
            "hazard_ratio": np.exp(params),
            "CI95_low": np.exp(params - z * bse),
            "CI95_high": np.exp(params + z * bse),
            "p_value": pvalues,
        }
    )
    qc = {
        "model": model_name,
        "n": len(fit_data),
        "events": int(fit_data["OS_status"].sum()),
        "all_coefficients_finite": bool(np.isfinite(params).all()),
        "all_confidence_intervals_finite": bool(
            np.isfinite(terms[["CI95_low", "CI95_high"]].to_numpy()).all()
        ),
        "maximum_hazard_ratio": float(terms["hazard_ratio"].max()),
        "score_hazard_ratio": float(terms.loc[terms["term"] == "score_z", "hazard_ratio"].iloc[0]),
        "score_CI95_low": float(terms.loc[terms["term"] == "score_z", "CI95_low"].iloc[0]),
        "score_CI95_high": float(terms.loc[terms["term"] == "score_z", "CI95_high"].iloc[0]),
        "score_p_value": float(terms.loc[terms["term"] == "score_z", "p_value"].iloc[0]),
    }
    return terms, qc


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        for column in ws.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 45)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--series-matrix", required=True, type=Path)
    parser.add_argument("--annotation", required=True, type=Path)
    parser.add_argument("--clinical", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    matrix = read_geo_table(
        args.series_matrix, "!series_matrix_table_begin", "!series_matrix_table_end"
    )
    annotation = parse_annotation(args.annotation)
    expression = extract_model_expression(matrix, annotation)
    analysis = build_analysis_table(args.clinical, expression)

    primary_covariates = ["AFP_high", "cirrhosis_yes", "multinodular_yes", "score_z"]
    size_covariates = [
        "AFP_high",
        "cirrhosis_yes",
        "size_large",
        "multinodular_yes",
        "score_z",
    ]
    primary_terms, primary_qc = fit_cox(analysis, "primary_stable_binary_model", primary_covariates)
    size_terms, size_qc = fit_cox(analysis, "with_size_stable_binary_model", size_covariates)
    terms = pd.concat([primary_terms, size_terms], ignore_index=True)
    qc = pd.DataFrame([primary_qc, size_qc])
    qc["separation_check_pass"] = (
        qc["all_coefficients_finite"]
        & qc["all_confidence_intervals_finite"]
        & (qc["maximum_hazard_ratio"] < 100)
    )

    notes = pd.DataFrame(
        [
            (
                "Stable AFP recoding",
                "AFP was represented by one binary indicator: high (>300 ng/mL) versus low (<=300 ng/mL).",
            ),
            (
                "Separation boundary",
                "Models are considered free of obvious separation when coefficients and confidence intervals are finite and no hazard ratio is extreme.",
            ),
            (
                "Interpretation",
                "These are retrospective external score-level context models, not clinical validation.",
            ),
        ],
        columns=["item", "note"],
    )

    analysis.to_csv(args.output_dir / "GSE14520_stable_cox_source_data.csv", index=False)
    terms.to_csv(args.output_dir / "GSE14520_stable_cox_terms.csv", index=False)
    qc.to_csv(args.output_dir / "GSE14520_stable_cox_QC.csv", index=False)
    workbook = args.output_dir / "Supplementary_Table_S11.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        terms.to_excel(writer, sheet_name="Stable_Cox_terms", index=False)
        qc.to_excel(writer, sheet_name="Model_QC", index=False)
        notes.to_excel(writer, sheet_name="Notes", index=False)
    format_workbook(workbook)

    print(qc.to_string(index=False))
    if not qc["separation_check_pass"].all():
        raise RuntimeError("Stable Cox separation QC failed.")
    print(f"wrote={workbook}")


if __name__ == "__main__":
    main()
