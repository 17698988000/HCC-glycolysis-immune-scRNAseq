from __future__ import annotations

import argparse
import gzip
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from scipy.sparse import coo_matrix
from scipy.stats import spearmanr, wilcoxon


GLYCOLYSIS_22 = [
    "HK1", "HK2", "GPI", "PFKL", "PFKP", "PFKM", "ALDOA", "ALDOB", "ALDOC",
    "TPI1", "GAPDH", "PGK1", "PGAM1", "ENO1", "ENO2", "PKM", "LDHA", "LDHB",
    "SLC2A1", "SLC2A3", "PFKFB3", "GCK",
]
READOUT_COEFFICIENTS = {
    "TPI1": 0.3041908,
    "ENO1": 0.9639654,
    "LDHA": 1.3404374,
    "SLC2A1": 0.2424239,
}
LEAVE_FOUR_OUT = [gene for gene in GLYCOLYSIS_22 if gene not in READOUT_COEFFICIENTS]
FEATURES = ["MIF", "SPP1", "four_gene_readout"]


def read_selected_matrix(
    path: Path,
    n_genes: int,
    n_barcodes: int,
    selected_cols: np.ndarray,
):
    # The full GSE189903 matrix has approximately 120 million nonzero entries.
    # Read it in chunks and retain only the prespecified target-cell columns.
    header_rows = 0
    with gzip.open(path, "rt") as handle:
        while True:
            line = handle.readline()
            if not line:
                raise ValueError("Matrix Market file ended before the dimension line.")
            header_rows += 1
            if line.startswith("%"):
                continue
            dimensions = [int(value) for value in line.split()]
            break
    if dimensions[:2] != [n_genes, n_barcodes]:
        raise ValueError(
            f"Matrix shape {tuple(dimensions[:2])} does not match genes/barcodes "
            f"({n_genes}, {n_barcodes})"
        )

    col_map = np.full(n_barcodes + 1, -1, dtype=np.int32)
    col_map[selected_cols + 1] = np.arange(len(selected_cols), dtype=np.int32)
    row_parts = []
    col_parts = []
    data_parts = []
    reader = pd.read_csv(
        path,
        sep=" ",
        compression="gzip",
        skiprows=header_rows,
        header=None,
        names=["row", "col", "value"],
        dtype=np.int32,
        chunksize=2_000_000,
    )
    for chunk in reader:
        matrix_cols = chunk["col"].to_numpy(copy=False)
        mapped_cols = col_map[matrix_cols]
        keep = mapped_cols >= 0
        if np.any(keep):
            row_parts.append(chunk["row"].to_numpy(copy=False)[keep] - 1)
            col_parts.append(mapped_cols[keep])
            data_parts.append(chunk["value"].to_numpy(copy=False)[keep])
    if not data_parts:
        raise ValueError("No nonzero counts were found for the selected cells.")
    return coo_matrix(
        (
            np.concatenate(data_parts).astype(np.int32, copy=False),
            (
                np.concatenate(row_parts).astype(np.int32, copy=False),
                np.concatenate(col_parts).astype(np.int32, copy=False),
            ),
        ),
        shape=(n_genes, len(selected_cols)),
    ).tocsc()


def load_inputs(data_dir: Path):
    info = pd.read_csv(data_dir / "GSE189903_Info.txt.gz", sep="\t", compression="gzip")
    barcodes = pd.read_csv(
        data_dir / "GSE189903_barcodes.tsv.gz",
        sep="\t",
        compression="gzip",
        header=None,
        names=["Cell"],
    )
    genes = pd.read_csv(
        data_dir / "GSE189903_genes.tsv.gz",
        sep="\t",
        compression="gzip",
        header=None,
        names=["ensembl_id", "gene_symbol"],
    )
    if not info["Cell"].equals(barcodes["Cell"]):
        raise ValueError("Metadata cell order does not match barcode order.")
    # GSE189903 contains HCC ("H") and intrahepatic cholangiocarcinoma ("C").
    # Restrict to HCC tumor-region samples such as 1HT1, 2HT2, and 4HT3.
    tumor_malignant_mask = info["Type"].eq("Malignant cell") & info["Sample"].str.match(
        r"^\d+HT\d+$"
    )
    selected_info = info.loc[tumor_malignant_mask].copy().reset_index(drop=True)
    selected_cols = np.flatnonzero(tumor_malignant_mask.to_numpy())
    selected_counts = read_selected_matrix(
        data_dir / "GSE189903_matrix.mtx.gz",
        len(genes),
        len(barcodes),
        selected_cols,
    )
    return selected_info, genes, selected_counts


def unique_gene_row_map(genes: pd.DataFrame) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, symbol in enumerate(genes["gene_symbol"].astype(str)):
        if symbol not in mapping:
            mapping[symbol] = idx
    return mapping


def normalize_log1p(selected_counts):
    selected_counts = selected_counts.tocsc().astype(np.float64)
    lib = np.asarray(selected_counts.sum(axis=0)).ravel()
    if np.any(lib <= 0):
        raise ValueError("Selected cells contain zero-library columns.")
    scale = 10000.0 / lib
    normalized = selected_counts.multiply(scale).tocsc()
    normalized.data = np.log1p(normalized.data)
    return normalized, lib


def rank_based_gene_set_score(normalized, gene_rows: list[int]) -> np.ndarray:
    # AUCell-like rank score: mean percentile rank of locked genes within each cell.
    # Deterministic gene-row ordering breaks expression ties without optimization.
    dense = normalized.toarray()
    n_genes, n_cells = dense.shape
    scores = np.empty(n_cells, dtype=float)
    gene_rows_arr = np.asarray(gene_rows, dtype=int)
    for j in range(n_cells):
        values = dense[:, j]
        order = np.lexsort((np.arange(n_genes), -values))
        ranks = np.empty(n_genes, dtype=np.int32)
        ranks[order] = np.arange(1, n_genes + 1)
        percentile = 1.0 - (ranks[gene_rows_arr] - 1) / max(n_genes - 1, 1)
        scores[j] = percentile.mean()
    return scores


def rank_balanced_groups(values: np.ndarray, cells: pd.Series) -> np.ndarray:
    order = np.lexsort((cells.astype(str).to_numpy(), values))
    groups = np.empty(len(values), dtype=object)
    low_n = int(np.ceil(len(values) / 2))
    groups[order[:low_n]] = "GlycoLow"
    groups[order[low_n:]] = "GlycoHigh"
    return groups


def patient_summaries(df: pd.DataFrame, score_col: str, group_col: str, score_name: str):
    correlation_rows = []
    effect_rows = []
    for patient, patient_df in df.groupby("patient", sort=True):
        for feature in FEATURES:
            valid_corr = patient_df[[score_col, feature]].dropna()
            if (
                len(valid_corr) < 3
                or valid_corr[score_col].nunique() < 2
                or valid_corr[feature].nunique() < 2
            ):
                rho, p = np.nan, np.nan
            else:
                rho, p = spearmanr(valid_corr[score_col], valid_corr[feature])
            correlation_rows.append(
                {
                    "score": score_name,
                    "patient": patient,
                    "feature": feature,
                    "n_cells": len(patient_df),
                    "spearman_rho": rho,
                    "spearman_p": p,
                }
            )
            means = patient_df.groupby(group_col)[feature].mean()
            effect_rows.append(
                {
                    "score": score_name,
                    "patient": patient,
                    "feature": feature,
                    "GlycoLow": means.get("GlycoLow", np.nan),
                    "GlycoHigh": means.get("GlycoHigh", np.nan),
                    "effect_High_minus_Low": means.get("GlycoHigh", np.nan)
                    - means.get("GlycoLow", np.nan),
                }
            )
    correlations = pd.DataFrame(correlation_rows)
    effects = pd.DataFrame(effect_rows)

    summary_rows = []
    for (score, feature), part in effects.groupby(["score", "feature"], sort=True):
        valid = part.dropna(subset=["GlycoLow", "GlycoHigh"])
        if len(valid) >= 2:
            test = wilcoxon(valid["GlycoHigh"], valid["GlycoLow"], alternative="two-sided")
            paired_p = float(test.pvalue)
        else:
            paired_p = np.nan
        corr_part = correlations[
            (correlations["score"] == score) & (correlations["feature"] == feature)
        ]
        summary_rows.append(
            {
                "score": score,
                "feature": feature,
                "n_patients": len(valid),
                "n_positive_effect": int((valid["effect_High_minus_Low"] > 0).sum()),
                "mean_effect_High_minus_Low": valid["effect_High_minus_Low"].mean(),
                "median_effect_High_minus_Low": valid["effect_High_minus_Low"].median(),
                "paired_wilcoxon_p": paired_p,
                "n_positive_spearman": int((corr_part["spearman_rho"] > 0).sum()),
                "median_spearman_rho": corr_part["spearman_rho"].median(),
            }
        )
    summary = pd.DataFrame(summary_rows)
    summary["paired_wilcoxon_FDR_within_score"] = np.nan
    # Three features per score; explicit BH implementation.
    for score, idx in summary.groupby("score").groups.items():
        pvals = summary.loc[idx, "paired_wilcoxon_p"].to_numpy(dtype=float)
        order = np.argsort(pvals)
        adjusted = np.empty_like(pvals)
        ranked = pvals[order] * len(pvals) / np.arange(1, len(pvals) + 1)
        ranked = np.minimum.accumulate(ranked[::-1])[::-1]
        adjusted[order] = np.minimum(ranked, 1.0)
        summary.loc[idx, "paired_wilcoxon_FDR_within_score"] = adjusted
    return correlations, effects, summary


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        for column in ws.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 48)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-dir", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    selected_info, genes, selected_counts = load_inputs(args.data_dir)
    gene_map = unique_gene_row_map(genes)
    required = set(GLYCOLYSIS_22 + list(READOUT_COEFFICIENTS) + ["MIF", "SPP1"])
    missing = sorted(required - set(gene_map))
    if missing:
        raise ValueError(f"Missing required genes: {missing}")

    selected_info["patient"] = "HCC" + selected_info["Sample"].str.extract(r"^(\d+)")[0]
    normalized, library_size = normalize_log1p(selected_counts)
    selected_info["library_size"] = library_size

    score_rows_22 = [gene_map[gene] for gene in GLYCOLYSIS_22]
    score_rows_18 = [gene_map[gene] for gene in LEAVE_FOUR_OUT]
    selected_info["glycolysis_rank_score_22"] = rank_based_gene_set_score(normalized, score_rows_22)
    selected_info["glycolysis_rank_score_leave_four_out"] = rank_based_gene_set_score(
        normalized, score_rows_18
    )
    selected_info["GlycoGroup_22"] = rank_balanced_groups(
        selected_info["glycolysis_rank_score_22"].to_numpy(), selected_info["Cell"]
    )
    selected_info["GlycoGroup_leave_four_out"] = rank_balanced_groups(
        selected_info["glycolysis_rank_score_leave_four_out"].to_numpy(), selected_info["Cell"]
    )

    for gene in list(READOUT_COEFFICIENTS) + ["MIF", "SPP1"]:
        row = gene_map[gene]
        selected_info[gene] = np.asarray(normalized[row, :].todense()).ravel()
    selected_info["four_gene_readout"] = sum(
        selected_info[gene] * coef for gene, coef in READOUT_COEFFICIENTS.items()
    )

    corr22, eff22, sum22 = patient_summaries(
        selected_info, "glycolysis_rank_score_22", "GlycoGroup_22", "locked_22_gene_rank_score"
    )
    corr18, eff18, sum18 = patient_summaries(
        selected_info,
        "glycolysis_rank_score_leave_four_out",
        "GlycoGroup_leave_four_out",
        "leave_four_out_18_gene_rank_score",
    )
    correlations = pd.concat([corr22, corr18], ignore_index=True)
    effects = pd.concat([eff22, eff18], ignore_index=True)
    summary = pd.concat([sum22, sum18], ignore_index=True)

    cohort_summary = (
        selected_info.groupby("patient", sort=True)
        .agg(
            n_tumor_malignant_cells=("Cell", "size"),
            n_samples=("Sample", "nunique"),
            samples=("Sample", lambda x: ";".join(sorted(set(x)))),
            GlycoHigh_22_n=("GlycoGroup_22", lambda x: int((x == "GlycoHigh").sum())),
            GlycoLow_22_n=("GlycoGroup_22", lambda x: int((x == "GlycoLow").sum())),
            GlycoHigh_leave_four_out_n=(
                "GlycoGroup_leave_four_out",
                lambda x: int((x == "GlycoHigh").sum()),
            ),
            GlycoLow_leave_four_out_n=(
                "GlycoGroup_leave_four_out",
                lambda x: int((x == "GlycoLow").sum()),
            ),
        )
        .reset_index()
    )
    overall = pd.DataFrame(
        [
            ("dataset", "GSE189903"),
            (
                "selection",
                "Author-annotated Malignant cell entries from HCC tumor-region samples matching ^[0-9]+HT[0-9]+$",
            ),
            ("n_cells", len(selected_info)),
            ("n_patients", selected_info["patient"].nunique()),
            ("n_tumor_samples", selected_info["Sample"].nunique()),
            (
                "grouping_rule",
                "Locked gene sets and deterministic rank-balanced median grouping; no re-optimization",
            ),
            (
                "boundary",
                "Independent public-cohort replication of transcriptomic associations; not functional validation",
            ),
        ],
        columns=["metric", "value"],
    )

    selected_info.to_csv(args.output_dir / "GSE189903_locked_replication_cell_source.csv", index=False)
    cohort_summary.to_csv(args.output_dir / "GSE189903_locked_replication_cohort_summary.csv", index=False)
    correlations.to_csv(args.output_dir / "GSE189903_locked_replication_patient_correlations.csv", index=False)
    effects.to_csv(args.output_dir / "GSE189903_locked_replication_patient_effects.csv", index=False)
    summary.to_csv(args.output_dir / "GSE189903_locked_replication_summary.csv", index=False)

    workbook = args.output_dir / "Supplementary_Table_S12.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        overall.to_excel(writer, sheet_name="Overall", index=False)
        cohort_summary.to_excel(writer, sheet_name="Cohort_summary", index=False)
        summary.to_excel(writer, sheet_name="Replication_summary", index=False)
        correlations.to_excel(writer, sheet_name="Patient_correlations", index=False)
        effects.to_excel(writer, sheet_name="Patient_effects", index=False)
    format_workbook(workbook)

    print(overall.to_string(index=False))
    print(summary.to_string(index=False))
    print(f"wrote={workbook}")


if __name__ == "__main__":
    main()
