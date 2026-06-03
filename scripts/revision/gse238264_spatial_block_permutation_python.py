from __future__ import annotations

import argparse
import gzip
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from scipy.io import mmread
from scipy.stats import rankdata


GLYCOLYSIS_GENES = [
    "HK1",
    "HK2",
    "GPI",
    "PFKL",
    "PFKP",
    "PFKM",
    "ALDOA",
    "ALDOB",
    "ALDOC",
    "TPI1",
    "GAPDH",
    "PGK1",
    "PGAM1",
    "ENO1",
    "ENO2",
    "PKM",
    "LDHA",
    "LDHB",
    "SLC2A1",
    "SLC2A3",
    "PFKFB3",
    "GCK",
]
READOUT_COEFFICIENTS = {
    "TPI1": 0.3041908,
    "ENO1": 0.9639654,
    "LDHA": 1.3404374,
    "SLC2A1": 0.2424239,
}
LEAVE_FOUR_OUT_GENES = [
    gene for gene in GLYCOLYSIS_GENES if gene not in READOUT_COEFFICIENTS
]
OUTCOMES = ["ENO1", "four_gene_score", "PTGES", "MIF", "MIF_SPP1_ligand_score", "SPP1"]
SAMPLES = ["HCC1R", "HCC2R", "HCC3R", "HCC4R"]


def read_gzip_lines(path: Path) -> list[str]:
    with gzip.open(path, "rt", encoding="utf-8", errors="strict") as handle:
        return [line.rstrip("\r\n") for line in handle]


def find_sample_dir(root: Path, sample: str) -> Path:
    candidates = [root / sample / sample, root / sample]
    for candidate in candidates:
        if (
            (candidate / "filtered_feature_bc_matrix" / "matrix.mtx.gz").exists()
            and (candidate / "spatial" / "tissue_positions_list.csv").exists()
        ):
            return candidate
    raise FileNotFoundError(f"Cannot locate extracted Visium directory for {sample}")


def bh_adjust(pvals: np.ndarray) -> np.ndarray:
    result = np.full(len(pvals), np.nan, dtype=float)
    valid = np.isfinite(pvals)
    if not valid.any():
        return result
    values = pvals[valid]
    order = np.argsort(values)
    ranked = values[order] * len(values) / np.arange(1, len(values) + 1)
    ranked = np.minimum.accumulate(ranked[::-1])[::-1]
    adjusted = np.empty_like(values)
    adjusted[order] = np.minimum(ranked, 1.0)
    result[valid] = adjusted
    return result


def empirical_spearman(
    x: np.ndarray, y: np.ndarray, rng: np.random.Generator, permutations: int
) -> tuple[float, float, float]:
    xr = rankdata(x).astype(float)
    yr = rankdata(y).astype(float)
    xr -= xr.mean()
    yr -= yr.mean()
    denom = np.sqrt(np.dot(xr, xr) * np.dot(yr, yr))
    if denom == 0:
        return np.nan, np.nan, np.nan
    observed = float(np.dot(xr, yr) / denom)
    more_positive = 0
    more_two_sided = 0
    for _ in range(permutations):
        permuted = rng.permutation(yr)
        rho = float(np.dot(xr, permuted) / denom)
        more_positive += rho >= observed
        more_two_sided += abs(rho) >= abs(observed)
    return (
        observed,
        (more_positive + 1) / (permutations + 1),
        (more_two_sided + 1) / (permutations + 1),
    )


def load_sample(sample_dir: Path, sample: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    matrix_dir = sample_dir / "filtered_feature_bc_matrix"
    matrix = mmread(matrix_dir / "matrix.mtx.gz").tocsr()
    barcodes = read_gzip_lines(matrix_dir / "barcodes.tsv.gz")
    feature_rows = [line.split("\t") for line in read_gzip_lines(matrix_dir / "features.tsv.gz")]
    genes = [row[1] if len(row) > 1 else row[0] for row in feature_rows]
    if matrix.shape != (len(genes), len(barcodes)):
        raise ValueError(f"{sample} matrix dimensions do not match features/barcodes")

    library_size = np.asarray(matrix.sum(axis=0)).ravel().astype(float)
    n_features = np.asarray((matrix > 0).sum(axis=0)).ravel().astype(int)
    mt_rows = [index for index, gene in enumerate(genes) if gene.upper().startswith("MT-")]
    mt_counts = (
        np.asarray(matrix[mt_rows, :].sum(axis=0)).ravel().astype(float)
        if mt_rows
        else np.zeros(len(barcodes), dtype=float)
    )
    percent_mt = np.divide(
        mt_counts * 100.0,
        library_size,
        out=np.zeros_like(mt_counts),
        where=library_size > 0,
    )
    keep = (n_features >= 200) & (percent_mt < 25) & (library_size > 0)

    positions = pd.read_csv(
        sample_dir / "spatial" / "tissue_positions_list.csv",
        header=None,
        names=["barcode", "in_tissue", "array_row", "array_col", "pixel_row", "pixel_col"],
    )
    positions = positions.set_index("barcode").reindex(barcodes)
    if positions[["array_row", "array_col"]].isna().any().any():
        raise ValueError(f"{sample} positions are missing for one or more matrix barcodes")

    gene_to_rows: dict[str, list[int]] = {}
    for index, gene in enumerate(genes):
        gene_to_rows.setdefault(gene, []).append(index)
    required = set(GLYCOLYSIS_GENES) | {"MIF", "SPP1", "PTGES"}
    missing = sorted(required - set(gene_to_rows))
    if missing:
        raise ValueError(f"{sample} missing required genes: {missing}")

    selected = pd.DataFrame(
        {
            "sample": sample,
            "barcode": np.asarray(barcodes)[keep],
            "library_size": library_size[keep],
            "n_features": n_features[keep],
            "percent_mt": percent_mt[keep],
            "array_row": positions["array_row"].to_numpy()[keep].astype(int),
            "array_col": positions["array_col"].to_numpy()[keep].astype(int),
        }
    )
    for gene in sorted(required):
        rows = gene_to_rows[gene]
        counts = np.asarray(matrix[rows, :].sum(axis=0)).ravel().astype(float)[keep]
        selected[gene] = np.log1p(counts / selected["library_size"].to_numpy() * 10000.0)

    selected["glycolysis_22_gene_mean_logexpr"] = selected[GLYCOLYSIS_GENES].mean(axis=1)
    selected["glycolysis_leave_four_out_18_gene_mean_logexpr"] = selected[
        LEAVE_FOUR_OUT_GENES
    ].mean(axis=1)
    selected["four_gene_score"] = sum(
        selected[gene] * coefficient for gene, coefficient in READOUT_COEFFICIENTS.items()
    )
    selected["MIF_SPP1_ligand_score"] = selected[["MIF", "SPP1"]].mean(axis=1)
    selected["block_row"] = selected["array_row"] // 6
    selected["block_col"] = selected["array_col"] // 6
    selected["block_id"] = (
        selected["block_row"].astype(str) + "_" + selected["block_col"].astype(str)
    )

    qc = pd.DataFrame(
        [
            ("sample", sample),
            ("matrix_barcodes", len(barcodes)),
            ("spots_after_qc", int(keep.sum())),
            ("min_features", 200),
            ("max_percent_mt", 25),
            ("block_definition", "fixed 6x6 array-coordinate bins"),
        ],
        columns=["metric", "value"],
    )
    return selected, qc


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-root", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--permutations", type=int, default=10000)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    spot_frames = []
    qc_frames = []
    for sample in SAMPLES:
        spot_df, qc_df = load_sample(find_sample_dir(args.data_root, sample), sample)
        spot_frames.append(spot_df)
        qc_frames.append(qc_df)
        print(f"loaded={sample} spots={len(spot_df)}")
    spots = pd.concat(spot_frames, ignore_index=True)
    source_qc = pd.concat(qc_frames, ignore_index=True)

    score_columns = {
        "locked_22_gene_mean_logexpr": "glycolysis_22_gene_mean_logexpr",
        "leave_four_out_18_gene_mean_logexpr": "glycolysis_leave_four_out_18_gene_mean_logexpr",
    }
    rng = np.random.default_rng(20260603)
    block_rows = []
    result_rows = []
    for sample, sample_df in spots.groupby("sample", sort=True):
        block_means = (
            sample_df.groupby("block_id", sort=True)
            .agg(
                n_spots=("barcode", "size"),
                array_row_mean=("array_row", "mean"),
                array_col_mean=("array_col", "mean"),
                **{
                    column: (column, "mean")
                    for column in list(score_columns.values()) + OUTCOMES
                },
            )
            .reset_index()
        )
        block_means = block_means[block_means["n_spots"] >= 10].copy()
        block_means.insert(0, "sample", sample)
        block_rows.append(block_means)
        for score_name, score_col in score_columns.items():
            for outcome in OUTCOMES:
                rho, p_positive, p_two = empirical_spearman(
                    block_means[score_col].to_numpy(dtype=float),
                    block_means[outcome].to_numpy(dtype=float),
                    rng,
                    args.permutations,
                )
                result_rows.append(
                    {
                        "sample": sample,
                        "score": score_name,
                        "outcome": outcome,
                        "n_spots": len(sample_df),
                        "n_spatial_blocks": len(block_means),
                        "block_size_rule": "fixed 6x6 array-coordinate bins; blocks with >=10 spots",
                        "spearman_rho_block_means": rho,
                        "empirical_p_positive": p_positive,
                        "empirical_p_two_sided": p_two,
                        "n_permutations": args.permutations,
                        "interpretation_boundary": (
                            "Spatial-block-aggregated empirical sensitivity; not RCTD-adjusted "
                            "and not evidence of direct signaling"
                        ),
                    }
                )
    blocks = pd.concat(block_rows, ignore_index=True)
    results = pd.DataFrame(result_rows)
    results["FDR_two_sided_global"] = bh_adjust(
        results["empirical_p_two_sided"].to_numpy(dtype=float)
    )
    results["FDR_positive_global"] = bh_adjust(
        results["empirical_p_positive"].to_numpy(dtype=float)
    )
    summary = (
        results.groupby(["score", "outcome"], sort=True)
        .agg(
            n_sections=("sample", "nunique"),
            n_positive_rho=("spearman_rho_block_means", lambda x: int((x > 0).sum())),
            median_rho=("spearman_rho_block_means", "median"),
            min_rho=("spearman_rho_block_means", "min"),
            max_rho=("spearman_rho_block_means", "max"),
            n_sections_empirical_p_positive_lt_0_05=(
                "empirical_p_positive",
                lambda x: int((x < 0.05).sum()),
            ),
            n_sections_empirical_p_two_sided_lt_0_05=(
                "empirical_p_two_sided",
                lambda x: int((x < 0.05).sum()),
            ),
        )
        .reset_index()
    )
    overall = pd.DataFrame(
        [
            ("dataset", "GSE238264"),
            ("sections", "HCC1R;HCC2R;HCC3R;HCC4R"),
            ("total_spots_after_qc", len(spots)),
            ("block_rule", "fixed 6x6 array-coordinate bins; retain blocks with >=10 spots"),
            ("permutations_per_test", args.permutations),
            (
                "boundary",
                "Independent spatial-block empirical sensitivity; not RCTD-adjusted and not direct signaling evidence",
            ),
        ],
        columns=["metric", "value"],
    )

    spots.to_csv(args.output_dir / "GSE238264_spatial_block_spot_source.csv", index=False)
    blocks.to_csv(args.output_dir / "GSE238264_spatial_block_means.csv", index=False)
    results.to_csv(args.output_dir / "GSE238264_spatial_block_permutation_results.csv", index=False)
    summary.to_csv(args.output_dir / "GSE238264_spatial_block_permutation_summary.csv", index=False)
    source_qc.to_csv(args.output_dir / "GSE238264_spatial_block_source_qc.csv", index=False)

    workbook = args.output_dir / "Supplementary_Table_S14.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        overall.to_excel(writer, sheet_name="Overall", index=False)
        source_qc.to_excel(writer, sheet_name="Source_QC", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        results.to_excel(writer, sheet_name="Section_results", index=False)
        blocks.to_excel(writer, sheet_name="Block_means", index=False)
    wb = load_workbook(workbook)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        for column in ws.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 48)
    wb.save(workbook)

    print(overall.to_string(index=False))
    print(summary.to_string(index=False))
    print(f"wrote={workbook}")


if __name__ == "__main__":
    main()
