from __future__ import annotations

import argparse
import gzip
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from scipy.stats import spearmanr, wilcoxon


GLYCOLYSIS_GENES = [
    "HK1",
    "HK2",
    "GPI",
    "PFKL",
    "PFKP",
    "PFKM",
    "ALDOA",
    "ALDOB",
    "ALDOC",
    "TPI1",
    "GAPDH",
    "PGK1",
    "PGAM1",
    "ENO1",
    "ENO2",
    "PKM",
    "LDHA",
    "LDHB",
    "SLC2A1",
    "SLC2A3",
    "PFKFB3",
    "GCK",
]
READOUT_COEFFICIENTS = {
    "TPI1": 0.3041908,
    "ENO1": 0.9639654,
    "LDHA": 1.3404374,
    "SLC2A1": 0.2424239,
}
LEAVE_FOUR_OUT_GENES = [
    gene for gene in GLYCOLYSIS_GENES if gene not in READOUT_COEFFICIENTS
]
FEATURES = ["MIF", "SPP1", "four_gene_readout"]


def rank_balanced_groups(values: np.ndarray, cells: pd.Series) -> np.ndarray:
    order = np.lexsort((cells.astype(str).to_numpy(), values))
    groups = np.empty(len(values), dtype=object)
    low_n = int(np.ceil(len(values) / 2))
    groups[order[:low_n]] = "GlycoLow"
    groups[order[low_n:]] = "GlycoHigh"
    return groups


def patient_internal_groups(df: pd.DataFrame, score_col: str) -> np.ndarray:
    groups = pd.Series(index=df.index, dtype=object)
    for _, part in df.groupby("patient", sort=True):
        groups.loc[part.index] = rank_balanced_groups(
            part[score_col].to_numpy(), part["cell"]
        )
    return groups.to_numpy()


def bh_adjust(pvals: np.ndarray) -> np.ndarray:
    result = np.full(len(pvals), np.nan, dtype=float)
    valid = np.isfinite(pvals)
    if not valid.any():
        return result
    values = pvals[valid]
    order = np.argsort(values)
    ranked = values[order] * len(values) / np.arange(1, len(values) + 1)
    ranked = np.minimum.accumulate(ranked[::-1])[::-1]
    adjusted = np.empty_like(values)
    adjusted[order] = np.minimum(ranked, 1.0)
    result[valid] = adjusted
    return result


def read_locked_cells(path: Path) -> pd.DataFrame:
    locked = pd.read_csv(path)
    required = {"cell", "patient", "Glycolysis_AUC", "GlycoGroup"}
    missing = required - set(locked.columns)
    if missing:
        raise ValueError(f"Locked source is missing columns: {sorted(missing)}")
    if locked["cell"].duplicated().any():
        raise ValueError("Locked source contains duplicate cell IDs")
    return locked[["cell", "patient", "Glycolysis_AUC", "GlycoGroup"]].copy()


def extract_locked_expression(count_path: Path, locked: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    required_genes = set(GLYCOLYSIS_GENES) | {"MIF", "SPP1"}
    with gzip.open(count_path, "rt", encoding="utf-8", errors="strict") as handle:
        matrix_cells = handle.readline().rstrip("\r\n").split("\t")
        cell_to_index = {cell: index for index, cell in enumerate(matrix_cells)}
        missing_cells = sorted(set(locked["cell"]) - set(cell_to_index))
        if missing_cells:
            raise ValueError(f"{len(missing_cells)} locked cells are missing from count matrix")
        selected_indices = np.asarray(
            [cell_to_index[cell] for cell in locked["cell"]], dtype=np.int64
        )
        library_size = np.zeros(len(locked), dtype=np.float64)
        gene_counts: dict[str, np.ndarray] = {}
        n_gene_rows = 0
        for line in handle:
            gene, values_text = line.rstrip("\r\n").split("\t", 1)
            values = np.fromstring(values_text, sep="\t", dtype=np.int32)
            if len(values) != len(matrix_cells):
                raise ValueError(
                    f"Count row length mismatch for {gene}: {len(values)} vs {len(matrix_cells)}"
                )
            selected = values[selected_indices].astype(np.float64)
            library_size += selected
            if gene in required_genes:
                gene_counts[gene] = gene_counts.get(gene, 0) + selected
            n_gene_rows += 1
            if n_gene_rows % 5000 == 0:
                print(f"parsed_gene_rows={n_gene_rows}")

    missing_genes = sorted(required_genes - set(gene_counts))
    if missing_genes:
        raise ValueError(f"Required genes are missing from count matrix: {missing_genes}")
    if np.any(library_size <= 0):
        raise ValueError("One or more locked cells have zero library size")

    expression = pd.DataFrame(
        {
            gene: np.log1p(counts / library_size * 10000.0)
            for gene, counts in gene_counts.items()
        }
    )
    expression.insert(0, "cell", locked["cell"].to_numpy())
    qc = pd.DataFrame(
        [
            ("count_matrix_cells", len(matrix_cells)),
            ("locked_cells", len(locked)),
            ("matched_locked_cells", len(locked)),
            ("count_matrix_gene_rows", n_gene_rows),
            ("required_genes_present", len(required_genes)),
            ("normalization", "log1p(raw_count / cell_library_size * 10000)"),
        ],
        columns=["metric", "value"],
    )
    return expression, qc


def patient_effects(df: pd.DataFrame, group_col: str, strategy: str) -> pd.DataFrame:
    rows = []
    for patient, part in df.groupby("patient", sort=True):
        for feature in FEATURES:
            means = part.groupby(group_col)[feature].mean()
            rows.append(
                {
                    "strategy": strategy,
                    "patient": patient,
                    "feature": feature,
                    "n_cells": len(part),
                    "GlycoLow": means.get("GlycoLow", np.nan),
                    "GlycoHigh": means.get("GlycoHigh", np.nan),
                    "effect_High_minus_Low": means.get("GlycoHigh", np.nan)
                    - means.get("GlycoLow", np.nan),
                }
            )
    return pd.DataFrame(rows)


def patient_correlations(df: pd.DataFrame, score_col: str, score_name: str) -> pd.DataFrame:
    rows = []
    for patient, part in df.groupby("patient", sort=True):
        for feature in FEATURES:
            valid = part[[score_col, feature]].dropna()
            if (
                len(valid) < 3
                or valid[score_col].nunique() < 2
                or valid[feature].nunique() < 2
            ):
                rho, p = np.nan, np.nan
            else:
                rho, p = spearmanr(valid[score_col], valid[feature])
            rows.append(
                {
                    "score": score_name,
                    "patient": patient,
                    "feature": feature,
                    "n_cells": len(valid),
                    "spearman_rho": rho,
                    "spearman_p": p,
                }
            )
    return pd.DataFrame(rows)


def summarize_effects(effects: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (strategy, feature), part in effects.groupby(["strategy", "feature"], sort=True):
        valid = part.dropna(subset=["GlycoLow", "GlycoHigh"])
        p = (
            float(wilcoxon(valid["GlycoHigh"], valid["GlycoLow"], alternative="two-sided").pvalue)
            if len(valid) >= 2
            else np.nan
        )
        rows.append(
            {
                "strategy": strategy,
                "feature": feature,
                "n_patients": len(valid),
                "n_positive_effect": int((valid["effect_High_minus_Low"] > 0).sum()),
                "mean_effect_High_minus_Low": valid["effect_High_minus_Low"].mean(),
                "median_effect_High_minus_Low": valid["effect_High_minus_Low"].median(),
                "paired_wilcoxon_p": p,
            }
        )
    summary = pd.DataFrame(rows)
    summary["paired_wilcoxon_FDR_within_strategy"] = np.nan
    for strategy, idx in summary.groupby("strategy").groups.items():
        summary.loc[idx, "paired_wilcoxon_FDR_within_strategy"] = bh_adjust(
            summary.loc[idx, "paired_wilcoxon_p"].to_numpy(dtype=float)
        )
    return summary


def summarize_correlations(correlations: pd.DataFrame) -> pd.DataFrame:
    return (
        correlations.groupby(["score", "feature"], sort=True)
        .agg(
            n_patients=("patient", "nunique"),
            n_positive_spearman=("spearman_rho", lambda x: int((x > 0).sum())),
            median_spearman_rho=("spearman_rho", "median"),
            min_spearman_rho=("spearman_rho", "min"),
            max_spearman_rho=("spearman_rho", "max"),
        )
        .reset_index()
    )


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        for column in ws.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 48)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--count-matrix", type=Path, required=True)
    parser.add_argument("--locked-cells", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    locked = read_locked_cells(args.locked_cells)
    expression, source_qc = extract_locked_expression(args.count_matrix, locked)
    df = locked.merge(expression, on="cell", validate="one_to_one")
    df["mean_logexpr_22_gene_score"] = df[GLYCOLYSIS_GENES].mean(axis=1)
    df["leave_four_out_18_gene_score"] = df[LEAVE_FOUR_OUT_GENES].mean(axis=1)
    df["four_gene_readout"] = sum(
        df[gene] * coefficient for gene, coefficient in READOUT_COEFFICIENTS.items()
    )
    df["leave_four_out_global_group"] = rank_balanced_groups(
        df["leave_four_out_18_gene_score"].to_numpy(), df["cell"]
    )
    df["leave_four_out_patient_internal_group"] = patient_internal_groups(
        df, "leave_four_out_18_gene_score"
    )

    score_qc_rows = [
        ("n_cells", len(df)),
        ("n_patients", df["patient"].nunique()),
        ("locked_original_GlycoHigh_n", int((df["GlycoGroup"] == "GlycoHigh").sum())),
        ("locked_original_GlycoLow_n", int((df["GlycoGroup"] == "GlycoLow").sum())),
        (
            "leave_four_out_global_concordance_with_locked_group",
            float((df["leave_four_out_global_group"] == df["GlycoGroup"]).mean()),
        ),
        (
            "leave_four_out_patient_internal_concordance_with_locked_group",
            float((df["leave_four_out_patient_internal_group"] == df["GlycoGroup"]).mean()),
        ),
        (
            "spearman_leave_four_out_vs_locked_AUCell",
            float(spearmanr(df["leave_four_out_18_gene_score"], df["Glycolysis_AUC"]).statistic),
        ),
        (
            "spearman_leave_four_out_vs_22_gene_mean_logexpr",
            float(
                spearmanr(
                    df["leave_four_out_18_gene_score"], df["mean_logexpr_22_gene_score"]
                ).statistic
            ),
        ),
        (
            "interpretation_boundary",
            "Expression-level locked-gene sensitivity; not a re-optimized replacement for the primary AUCell state definition",
        ),
    ]
    score_qc = pd.DataFrame(score_qc_rows, columns=["metric", "value"])

    effects = pd.concat(
        [
            patient_effects(df, "GlycoGroup", "locked_original_GlycoGroup"),
            patient_effects(
                df,
                "leave_four_out_global_group",
                "leave_four_out_global_rank_balanced",
            ),
            patient_effects(
                df,
                "leave_four_out_patient_internal_group",
                "leave_four_out_patient_internal_rank_balanced",
            ),
        ],
        ignore_index=True,
    )
    correlations = pd.concat(
        [
            patient_correlations(
                df, "Glycolysis_AUC", "locked_original_Glycolysis_AUC"
            ),
            patient_correlations(
                df,
                "leave_four_out_18_gene_score",
                "leave_four_out_18_gene_mean_logexpr_score",
            ),
        ],
        ignore_index=True,
    )
    effect_summary = summarize_effects(effects)
    correlation_summary = summarize_correlations(correlations)

    cell_columns = [
        "cell",
        "patient",
        "Glycolysis_AUC",
        "GlycoGroup",
        "mean_logexpr_22_gene_score",
        "leave_four_out_18_gene_score",
        "leave_four_out_global_group",
        "leave_four_out_patient_internal_group",
        "MIF",
        "SPP1",
        "four_gene_readout",
    ]
    df[cell_columns].to_csv(args.output_dir / "GSE149614_leave_four_out_cell_source.csv", index=False)
    source_qc.to_csv(args.output_dir / "GSE149614_leave_four_out_source_qc.csv", index=False)
    score_qc.to_csv(args.output_dir / "GSE149614_leave_four_out_score_qc.csv", index=False)
    effects.to_csv(args.output_dir / "GSE149614_leave_four_out_patient_effects.csv", index=False)
    correlations.to_csv(
        args.output_dir / "GSE149614_leave_four_out_patient_correlations.csv", index=False
    )
    effect_summary.to_csv(
        args.output_dir / "GSE149614_leave_four_out_effect_summary.csv", index=False
    )
    correlation_summary.to_csv(
        args.output_dir / "GSE149614_leave_four_out_correlation_summary.csv", index=False
    )

    workbook = args.output_dir / "Supplementary_Table_S13.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        source_qc.to_excel(writer, sheet_name="Source_QC", index=False)
        score_qc.to_excel(writer, sheet_name="Score_QC", index=False)
        effect_summary.to_excel(writer, sheet_name="Effect_summary", index=False)
        correlation_summary.to_excel(writer, sheet_name="Correlation_summary", index=False)
        effects.to_excel(writer, sheet_name="Patient_effects", index=False)
        correlations.to_excel(writer, sheet_name="Patient_correlations", index=False)
    format_workbook(workbook)

    print(score_qc.to_string(index=False))
    print(effect_summary.to_string(index=False))
    print(correlation_summary.to_string(index=False))
    print(f"wrote={workbook}")


if __name__ == "__main__":
    main()
