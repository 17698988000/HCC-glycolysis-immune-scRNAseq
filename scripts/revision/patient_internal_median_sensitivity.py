from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


REQUIRED_COLUMNS = {"cell", "patient", "Glycolysis_AUC", "GlycoGroup"}


def rank_balanced_within_patient_groups(df: pd.DataFrame) -> pd.DataFrame:
    out = df.sort_values(
        ["patient", "Glycolysis_AUC", "cell"], kind="mergesort"
    ).copy()
    out["patient_rank"] = out.groupby("patient").cumcount() + 1
    out["patient_n"] = out.groupby("patient")["cell"].transform("size")
    out["patient_internal_group"] = np.where(
        out["patient_rank"] <= np.ceil(out["patient_n"] / 2),
        "GlycoLow",
        "GlycoHigh",
    )
    out["concordant_with_global_assignment"] = (
        out["GlycoGroup"] == out["patient_internal_group"]
    )
    return out


def build_summaries(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    patient_summary = (
        df.groupby("patient", sort=True)
        .agg(
            n_cells=("cell", "size"),
            global_high_n=("GlycoGroup", lambda x: int((x == "GlycoHigh").sum())),
            global_low_n=("GlycoGroup", lambda x: int((x == "GlycoLow").sum())),
            patient_internal_high_n=(
                "patient_internal_group",
                lambda x: int((x == "GlycoHigh").sum()),
            ),
            patient_internal_low_n=(
                "patient_internal_group",
                lambda x: int((x == "GlycoLow").sum()),
            ),
            global_assignment_concordance=(
                "concordant_with_global_assignment",
                "mean",
            ),
            patient_median_Glycolysis_AUC=("Glycolysis_AUC", "median"),
        )
        .reset_index()
    )
    patient_summary["global_high_fraction"] = (
        patient_summary["global_high_n"] / patient_summary["n_cells"]
    )
    patient_summary["patient_internal_high_fraction"] = (
        patient_summary["patient_internal_high_n"] / patient_summary["n_cells"]
    )
    patient_summary = patient_summary[
        [
            "patient",
            "n_cells",
            "global_high_n",
            "global_low_n",
            "global_high_fraction",
            "patient_internal_high_n",
            "patient_internal_low_n",
            "patient_internal_high_fraction",
            "global_assignment_concordance",
            "patient_median_Glycolysis_AUC",
        ]
    ]

    contingency = pd.crosstab(
        df["GlycoGroup"],
        df["patient_internal_group"],
        rownames=["Global rank-balanced assignment"],
        colnames=["Patient-internal rank-balanced assignment"],
    ).reset_index()

    overall_summary = pd.DataFrame(
        [
            ("tumor_hepatocyte_n", len(df)),
            ("n_patients", df["patient"].nunique()),
            (
                "overall_assignment_concordance",
                df["concordant_with_global_assignment"].mean(),
            ),
            (
                "minimum_patient_concordance",
                patient_summary["global_assignment_concordance"].min(),
            ),
            (
                "maximum_patient_concordance",
                patient_summary["global_assignment_concordance"].max(),
            ),
            (
                "minimum_global_high_fraction",
                patient_summary["global_high_fraction"].min(),
            ),
            (
                "maximum_global_high_fraction",
                patient_summary["global_high_fraction"].max(),
            ),
            (
                "patients_with_global_high_majority",
                int((patient_summary["global_high_fraction"] > 0.5).sum()),
            ),
            (
                "patients_with_global_low_majority",
                int((patient_summary["global_high_fraction"] < 0.5).sum()),
            ),
            (
                "interpretation",
                (
                    "The global and patient-internal assignments were broadly concordant, "
                    "but global GlycoHigh proportions varied substantially across patients. "
                    "This supports a shared glycolysis gradient while confirming that "
                    "downstream inference should remain patient-aware."
                ),
            ),
            (
                "boundary",
                (
                    "This is a grouping-stability analysis only. Ligand expression, "
                    "communication, and readout associations were not recomputed because "
                    "the frozen release does not redistribute the full normalized matrix."
                ),
            ),
        ],
        columns=["metric", "value"],
    )
    return patient_summary, overall_summary, contingency


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for column in ws.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 55)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Patient-internal rank-balanced GlycoHigh/GlycoLow grouping sensitivity."
    )
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    df = pd.read_csv(args.input)
    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")
    if len(df) != 15391:
        raise ValueError(f"Expected 15,391 cells, observed {len(df):,}")

    assignments = rank_balanced_within_patient_groups(df)
    patient_summary, overall_summary, contingency = build_summaries(assignments)

    assignments.to_csv(
        args.output_dir / "patient_internal_median_cell_assignments.csv", index=False
    )
    patient_summary.to_csv(
        args.output_dir / "patient_internal_median_patient_summary.csv", index=False
    )
    overall_summary.to_csv(
        args.output_dir / "patient_internal_median_overall_summary.csv", index=False
    )
    contingency.to_csv(
        args.output_dir / "patient_internal_median_contingency.csv", index=False
    )

    workbook = args.output_dir / "Supplementary_Table_S10.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        patient_summary.to_excel(writer, sheet_name="Patient_summary", index=False)
        overall_summary.to_excel(writer, sheet_name="Overall_summary", index=False)
        contingency.to_excel(writer, sheet_name="Contingency", index=False)
    format_workbook(workbook)

    print(f"overall_assignment_concordance={assignments['concordant_with_global_assignment'].mean():.6f}")
    print(
        "global_high_fraction_range="
        f"{patient_summary['global_high_fraction'].min():.6f}-"
        f"{patient_summary['global_high_fraction'].max():.6f}"
    )
    print(f"wrote={workbook}")


if __name__ == "__main__":
    main()
